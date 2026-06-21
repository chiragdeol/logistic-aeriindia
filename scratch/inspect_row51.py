import openpyxl

path = r"c:\Users\Boxinall\Downloads\Logistic-main (1)\Logistic-main\USA SAVER Revised_allin 08-04-26.xlsx"
wb = openpyxl.load_workbook(path, data_only=False)

for sname in wb.sheetnames:
    print(f"\nSheet: {sname}")
    sheet = wb[sname]
    # Row 51 in Excel (which corresponds to row index 51, weight 30+)
    row_vals = [sheet.cell(row=51, column=c).value for c in range(1, 10)]
    print(f"Row 51: {row_vals}")
