import openpyxl

path = r"c:\Users\Boxinall\Downloads\Logistic-main (1)\Logistic-main\USA SAVER Revised_allin 08-04-26.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
sheet = wb["COST_UPS SAVER (Allin)"]

print("Row index, weight, USA value, number_format:")
for r in range(9, 60):
    w = sheet.cell(row=r, column=1).value
    val = sheet.cell(row=r, column=5).value
    fmt = sheet.cell(row=r, column=5).number_format
    print(f"Row {r}: weight={w}, USA={val}, format={fmt}")
