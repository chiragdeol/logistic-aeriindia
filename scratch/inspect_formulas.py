import openpyxl

path = r"c:\Users\Boxinall\Downloads\Logistic-main (1)\Logistic-main\USA SAVER Revised_allin 08-04-26.xlsx"
wb = openpyxl.load_workbook(path, data_only=False)

for sname in wb.sheetnames:
    print(f"\nSheet: {sname}")
    sheet = wb[sname]
    for r in range(1, 10):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
        print(f"Row {r}: {row_vals}")
